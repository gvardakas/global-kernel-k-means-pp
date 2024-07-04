import numpy as np
import pandas as pd
import os
import csv
from openpyxl import load_workbook, Workbook
from Common_Modules.Evaluation import Evaluator

class General_Functions:

    def __init__(self):
        self.evaluator = Evaluator()

    def create_and_load_workbook(self, workbook_path):
        if os.path.exists(workbook_path)!=True:
            workbook = Workbook()
            workbook.save(workbook_path)
        return load_workbook(workbook_path)
    
    def write_data(self, workbook_path, df):
        experiment_name, workbook = self.set_experiment_name(workbook_path)
        
        writer = pd.ExcelWriter(workbook_path, engine = 'openpyxl')
        writer.book = workbook
        
        df.to_excel(writer, sheet_name=experiment_name)

        writer.save()
        writer.close()

    def set_experiment_name(self, workbook_path):
        workbook = self.create_and_load_workbook(workbook_path)
        sheet_name_splitted_array = workbook.sheetnames[-1].split("_")
        
        experiment_name = ""
        if len(sheet_name_splitted_array) == 1:
            sheet_to_delete = workbook.get_sheet_by_name(sheet_name_splitted_array[0])
            workbook.remove_sheet(sheet_to_delete)
            experiment_name = "Experiment_1"
        else:
            experiment_name = "Experiment_" + str(int(sheet_name_splitted_array[1]) + 1)
        
        return experiment_name, workbook
    
    def get_experiment_name(self, workbook_path):
        workbook = self.create_and_load_workbook(workbook_path)
        return workbook.sheetnames[-1]
       
    def save_excel(self, data_dir_path, df):
        self.create_directory(data_dir_path)
        workbook_path = data_dir_path + "/Data.xlsx"
        self.write_data(workbook_path, df)
    
    def create_csv(inertias_, n_iters_, execution_times, file_path):

        # Ensure all keys are aligned
        keys = sorted(set(inertias_.keys()).union(set(n_iters_.keys())).union(set(execution_times.keys())))

        # Combine dictionaries
        combined_data = {
            "K": keys,
            "MSE": [inertias_.get(k, None) for k in keys],
            "ITERATIONS": [n_iters_.get(k, None) for k in keys],
            "EXECUTION TIME": [execution_times.get(k, None) for k in keys]
        }

        # Write the combined dictionary to CSV
        with open(file_path, 'w', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=combined_data.keys())

            # Write the header
            writer.writeheader()

            # Write the data
            for i in range(len(keys)):
                row = {key: combined_data[key][i] for key in combined_data}
                writer.writerow(row)
                
    def append_to_csv(file_path, new_row):
        
        # Ensure new_row is a dictionary with the appropriate keys
        expected_keys = ["K", "MSE", "ITERATIONS", "EXECUTION TIME"]
        if not all(key in new_row for key in expected_keys):
            raise ValueError(f"New row must contain the keys: {expected_keys}")

        # Append the new row to the CSV
        with open(file_path, 'a', newline='') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=expected_keys)
        
            # Write the new row
            writer.writerow(new_row)

    def append_to_file(self, labels_true, labels_predict, inertia_, execution_times_, file_path):
        total_execution_time = sum(execution_times_.values())
        acc, pur, nmi, ari = self.evaluator.evaluate_model(labels_true, labels_predict)
        evaluation_results = f"MSE: {inertia_:.2f} ACC: {acc:.2f} PUR: {pur:.2f} NMI: {nmi:.2f} ARI: {ari:.2f} ET(s): {total_execution_time:.2f}"
        print(evaluation_results)
        with open(file_path, 'a') as file:
            file.write(evaluation_results + '\n')

    def find_images_to_plot(self, min_label, max_label, data, similarity, labels):
        images = []
        for i in range(min_label, max_label):
            image_indices_and_max_similarities = []
            
            valid_indices = np.where(labels == i)[0]
            for index in valid_indices:
                cur_max_similarity = similarity[index][np.argmax(similarity[index])]
                image_indices_and_max_similarities.append((index, cur_max_similarity))

            sorted_image_indices_and_max_similarities = sorted(image_indices_and_max_similarities, key=lambda x: x[1], reverse=True)
            desired_similarity = sorted_image_indices_and_max_similarities[0][1]
            
            for index, max_similarity in sorted_image_indices_and_max_similarities:
                if(max_similarity <= desired_similarity):
                    images.append(data[index])
                    desired_similarity = desired_similarity - 0.1 * desired_similarity
                if(len(images) % 10 == 0):
                    break
        
        return images          
        
    def create_directory(self, directory_path):
        if not os.path.exists(directory_path):
            os.makedirs(directory_path)
            print(f"Directory '{directory_path}' created successfully.")
        else:
            print(f"Directory '{directory_path}' already exists.")    
    
